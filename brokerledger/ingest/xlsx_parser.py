"""XLSX bank-statement parser. Reuses CSV heuristics after extracting rows."""
from __future__ import annotations

import csv
import io
from pathlib import Path

from openpyxl import load_workbook

from .csv_parser import parse_csv
from .normalize import RawTransaction


def parse_xlsx(path: Path) -> list[RawTransaction]:
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        ws = wb.active
        buf = io.StringIO()
        writer = csv.writer(buf)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(["" if v is None else str(v) for v in row])
    finally:
        wb.close()
    # Write to a temp path and reuse CSV parser — simplest and most consistent.
    tmp = path.with_suffix(".__bl__.csv")
    try:
        tmp.write_text(buf.getvalue(), encoding="utf-8")
        return parse_csv(tmp)
    finally:
        if tmp.exists():
            tmp.unlink()


__all__ = ["parse_xlsx", "RawTransaction"]

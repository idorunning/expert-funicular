"""XLSX export builder — Transactions, Category Totals, Affordability, Audit."""
from __future__ import annotations

from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select

from ..affordability.calculator import AffordabilityReport, compute_for_client
from ..auth.session import get_current
from ..categorize.taxonomy import COMMITTED_CATEGORIES, DISCRETIONARY_CATEGORIES
from ..db.engine import session_scope
from ..db.models import Client, Statement, Transaction, User


_BOLD = Font(bold=True)
_HEADER_FILL = PatternFill(start_color="E0E7EF", end_color="E0E7EF", fill_type="solid")


def _autosize(ws, max_col: int) -> None:
    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        max_len = 8
        for cell in ws[letter]:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = min(len(val), 60)
        ws.column_dimensions[letter].width = max_len + 2


def _write_header(ws, headers: list[str]) -> None:
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _BOLD
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="left")
    ws.freeze_panes = "A2"


def _transactions_sheet(wb: Workbook, client_id: int) -> None:
    from ..categorize.flags import deserialize_flags, flag_display_name

    ws = wb.create_sheet("Transactions")
    headers = [
        "Date", "Description", "Merchant", "Amount (GBP)", "Direction",
        "Category", "Group", "Certainty", "Method", "Flags", "Flagged",
    ]
    _write_header(ws, headers)
    with session_scope() as s:
        rows = s.execute(
            select(Transaction).where(Transaction.client_id == client_id)
            .order_by(Transaction.posted_date.asc(), Transaction.id.asc())
        ).scalars().all()
        for r in rows:
            ws.append([
                r.posted_date,
                r.description_raw,
                r.merchant_normalized,
                float(r.amount) if r.amount is not None else None,
                r.direction,
                r.category,
                r.category_group,
                round(r.confidence, 3) if r.confidence is not None else None,
                r.source,
                ", ".join(flag_display_name(f) for f in deserialize_flags(r.flags)),
                "Yes" if r.needs_review else "",
            ])
    ws.auto_filter.ref = ws.dimensions
    _autosize(ws, len(headers))


def _category_totals_sheet(wb: Workbook, report: AffordabilityReport) -> None:
    ws = wb.create_sheet("Category Totals")
    headers = ["Category", "Group", "Count", "Total (GBP)", "Monthly avg (GBP)"]
    _write_header(ws, headers)
    months = max(report.months_in_window, 1.0)
    for cat in list(COMMITTED_CATEGORIES) + list(DISCRETIONARY_CATEGORIES):
        t = report.per_category.get(cat)
        if t is None:
            ws.append([cat, "", 0, 0.0, 0.0])
            continue
        ws.append([
            t.category,
            t.group,
            t.count,
            float(t.total),
            float((t.total / Decimal(str(months))).quantize(Decimal("0.01"))),
        ])
    _autosize(ws, len(headers))


def _affordability_sheet(wb: Workbook, report: AffordabilityReport, client_name: str) -> None:
    ws = wb.create_sheet("Affordability Summary")
    ws.append(["Affordability Summary"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    rows: list[tuple[str, object]] = [
        ("Client", client_name),
        ("Period start", report.period_start.isoformat() if report.period_start else ""),
        ("Period end", report.period_end.isoformat() if report.period_end else ""),
        ("Months in window", round(report.months_in_window, 2)),
        ("", ""),
        ("Detected income (total)", float(report.income_total)),
        ("Declared income (override)",
         float(report.declared_income) if report.declared_income is not None else ""),
        ("Committed expenditure (total)", float(report.committed_total)),
        ("Discretionary expenditure (total)", float(report.discretionary_total)),
        ("Outgoings total", float(report.outgoings_total)),
        ("Net disposable (income - outgoings)", float(report.net_disposable)),
        ("", ""),
        ("Monthly income", float(report.monthly_income)),
        ("Monthly committed", float(report.monthly_committed)),
        ("Monthly discretionary", float(report.monthly_discretionary)),
        ("Monthly net disposable", float(report.monthly_net_disposable)),
    ]
    for label, value in rows:
        ws.append([label, value])
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 22


def _audit_sheet(wb: Workbook, client: Client, user_label: str) -> None:
    ws = wb.create_sheet("Audit")
    ws.append(["Client", client.display_name])
    ws.append(["Reference", client.reference or ""])
    ws.append(["Exported by", user_label])
    ws.append(["Exported at", datetime.now(timezone.utc).isoformat()])
    ws.append([])
    headers = [
        "File",
        "SHA-256",
        "Kind",
        "Imported at",
        "Verified by",
        "Verified at",
    ]
    start_row = 6
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.font = _BOLD
        cell.fill = _HEADER_FILL
    with session_scope() as s:
        stmts = s.execute(
            select(Statement)
            .where(Statement.client_id == client.id)
            .order_by(Statement.imported_at.asc())
        ).scalars().all()
        verifier_cache: dict[int, str] = {}
        for st in stmts:
            verifier = ""
            if st.verified_by is not None:
                verifier = verifier_cache.get(st.verified_by) or ""
                if not verifier:
                    u = s.get(User, st.verified_by)
                    verifier = u.username if u else f"user#{st.verified_by}"
                    verifier_cache[st.verified_by] = verifier
            ws.append([
                st.original_name,
                st.file_sha256,
                st.file_kind,
                st.imported_at.isoformat(),
                verifier,
                st.verified_at.isoformat() if st.verified_at else "",
            ])
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 72
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 26
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 26


def export_client_workbook(client_id: int, out_path: Path,
                           declared_income: Decimal | None = None) -> Path:
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    report = compute_for_client(client_id, declared_income=declared_income)
    with session_scope() as s:
        client = s.get(Client, client_id)
        if client is None:
            raise ValueError(f"Client {client_id} not found")
        current = get_current()
        user_label = ""
        if current is not None:
            u = s.get(User, current.id)
            user_label = f"{u.username} ({u.role})" if u else current.username

    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)
    _transactions_sheet(wb, client_id)
    _category_totals_sheet(wb, report)
    _affordability_sheet(wb, report, client.display_name)
    _audit_sheet(wb, client, user_label)
    wb.save(str(out_path))
    return out_path

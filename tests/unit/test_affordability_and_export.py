from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from brokerledger.affordability.calculator import compute_for_client
from brokerledger.categorize.categorizer import categorize_statement
from brokerledger.categorize.llm_client import FakeLLMClient
from brokerledger.clients.service import create_client
from brokerledger.export.xlsx import export_client_workbook
from brokerledger.ingest.router import ingest_statement


def _write_csv(tmp_path: Path) -> Path:
    p = tmp_path / "statement.csv"
    p.write_text(
        "Date,Description,Debit,Credit,Balance\n"
        "01/03/2025,SALARY ACME LTD,,3500.00,3500.00\n"
        "02/03/2025,COUNCIL TAX,180.00,,3320.00\n"
        "05/03/2025,OCTOPUS ENERGY,120.00,,3200.00\n"
        "07/03/2025,NETFLIX,10.99,,3189.01\n"
        "10/03/2025,TESCO STORES,54.30,,3134.71\n"
        "01/04/2025,SALARY ACME LTD,,3500.00,6634.71\n",
        encoding="utf-8",
    )
    return p


def test_affordability_totals(logged_in_admin, tmp_path: Path):
    client = create_client("Afford Client")
    result = ingest_statement(client.id, _write_csv(tmp_path))
    categorize_statement(result.statement_id, llm=FakeLLMClient())

    report = compute_for_client(client.id)
    assert report.income_total == Decimal("7000.00")
    assert report.committed_total == Decimal("300.00")  # Council 180 + Energy 120
    assert report.discretionary_total == Decimal("65.29")  # Tesco 54.30 + Netflix 10.99
    assert report.outgoings_total == Decimal("365.29")
    assert report.net_disposable == Decimal("6634.71")


def test_xlsx_export_sheets(logged_in_admin, tmp_path: Path):
    client = create_client("Export Client")
    result = ingest_statement(client.id, _write_csv(tmp_path))
    categorize_statement(result.statement_id, llm=FakeLLMClient())

    out = tmp_path / "export.xlsx"
    export_client_workbook(client.id, out)
    assert out.exists()
    wb = load_workbook(out)
    assert set(wb.sheetnames) == {"Transactions", "Category Totals", "Affordability Summary", "Audit"}
    ws = wb["Transactions"]
    # header + 6 rows (opening balance is skipped because its amount is blank)
    assert ws.max_row >= 6
    # Spot-check Affordability Summary has income line
    afford = wb["Affordability Summary"]
    labels = [str(row[0].value or "") for row in afford.iter_rows()]
    assert any("Detected income (total)" in l for l in labels)
    assert any("Net disposable" in l for l in labels)
